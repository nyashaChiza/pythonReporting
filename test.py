from openpyxl import load_workbook
def get_data():
    path = 'static/uploads/NAC PAYE.xlsx'
    wb = load_workbook(path)
    ws = wb.active
    AidsUSD =0
    AidsLevy = 0
    TaxUSD = 0
    GrossIncomeUSD = 0
    totalFringeB = 0
    AidsZWL =0
    PAYETax= 0
    PAYETaxZWL=0
    totalEarningsUSD=0
    totalEarningsZWL = 0
    cols = [6,8,9,10,11,12,13,14,15,16]
    for i in cols:
        for x in (range(6,ws.max_row+1)):
            
            if i == 6:
                
                try:
                    print('testing..')
                    #print('AidsUSD: ',AidsUSD)
                    print('VALUE',ws.cell(row=x, column=i).value)
                    AidsUSD = AidsUSD + ws.cell(row=x, Column=i).value
                    
                    
                except:
                        continue
            if i == 8:
                try:
                    GrossIncomeUSD = GrossIncomeUSD + ws.cell(row=x, Column=i).value
                    
                except:
                        continue
            if i == 9:
                try:
                    TaxUSD = TaxUSD + ws.cell(row=x, Column=i).value
                    
                except:
                        continue
            if i == 10:
                try:
                    AidsZWL = AidsZWL + ws.cell(row=x, Column=i).value
                except:
                        continue

            if i == 11:
                try:
                    PAYETax = PAYETax + ws.cell(row=x, Column=i).value
                except:
                        continue
            if i == 12:
                try:
                    totalEarningsUSD = totalEarningsUSD + ws.cell(row=x, Column=i).value
                except:
                        continue
            if i == 13:
                try:
                    AidsLevy = AidsLevy + ws.cell(row=x, Column=i).value
                except:
                        continue
            if i == 14:
                try:
                    PAYETaxZWL = PAYETaxZWL + ws.cell(row=x, Column=i).value
                except:
                        continue

            if i == 15:
                try:
                    totalEarningsZWL = totalEarningsZWL + ws.cell(row=x, Column=i).value
                except:
                        continue

            if i == 16:
                try:
                    totalFringeB = totalFringeB + ws.cell(row=x, Column=i).value
                except:
                        continue

#--------------------------------------------------
    values = {}
    values['tr1'] = '{:.2f}'.format((totalEarningsUSD+totalEarningsZWL+totalFringeB)/2)
    values['tr2'] =  '{:.2f}'.format((totalEarningsZWL+totalFringeB)/2)
    values['tr3'] = '{:.2f}'.format((totalEarningsUSD)/2)
    values['tr4'] = '{:.2f}'.format((GrossIncomeUSD)/2)
    values['ne'] = ws.max_row-6
    values['gp1'] = '{:.2f}'.format((PAYETax+ PAYETaxZWL)/2)
    values['gp2'] = '{:.2f}'.format((PAYETaxZWL)/2)
    values['gp3'] = '{:.2f}'.format((PAYETax)/2)
    values['gp4'] = '{:.2f}'.format((TaxUSD)/2)
    values['al1'] = '{:.2f}'.format((AidsZWL+AidsLevy)/2)
    values['al2'] = '{:.2f}'.format((AidsLevy)/2)
    values['al3'] = '{:.2f}'.format((AidsZWL)/2)
    values['al4'] = '{:.2f}'.format((AidsUSD)/2)
    values['tt1'] = '{:.2f}'.format((PAYETax+PAYETaxZWL+AidsZWL+AidsLevy)/2)
    values['tt2'] = '{:.2f}'.format((PAYETaxZWL+AidsLevy)/2)
    values['tt3'] = '{:.2f}'.format((PAYETax+AidsZWL)/2)
    values['tt4'] = '{:.2f}'.format((TaxUSD+AidsUSD)/2)
    return values

print(get_data())