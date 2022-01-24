from flask import Flask, render_template, request, url_for, session, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from mailmerge import MailMerge
import os
from openpyxl import load_workbook
import json
from datetime import date
from sqlalchemy import DATE,func

#---------------------------------------------------------------------------------------------------------------------
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] ='postgresql://NyashaChiza:Ch!zampeni95@petalmsql.postgres.database.azure.com/postgres?sslmode=require'
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = 'this_is_my_secret_key'
db = SQLAlchemy(app)
#---------------------------------------------------------------------------------------------------------------------

class A_I(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    sin = db.Column(db.String(10), nullable=False)
    gender = db.Column(db.String(1), nullable=False)
    first_name = db.Column(db.String(30), nullable=False)
    last_name = db.Column(db.String(30), nullable=False)
    last_name_prefix = db.Column(db.String(30), nullable=False)    
    academic_title = db.Column(db.String(30), nullable=False)    
    filler1 = db.Column(db.String(30), nullable=False)    
    birth_date = db.Column(db.String(30), nullable=False)    
    date_entry = db.Column(db.Date,  default=date.today())
class J_S(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    country_code = db.Column(db.String(2))
    postal_code = db.Column(db.String(11))
    city = db.Column(db.String(30))
    c_o = db.Column(db.String(40))
    filler2 = db.Column(db.String(30))
    filler3 = db.Column(db.String(30))
    filler4 = db.Column(db.String(30))
    country_code_mail_address = db.Column(db.String(2))    
    postal_code_mail_address = db.Column(db.String(11))    
    city_mail_address = db.Column(db.String(30))
    date_entry = db.Column(db.Date,  default=date.today())
class T_AC(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    date_of_entry_into_group  = db.Column(db.String(30))
    date_of_leaving_group = db.Column(db.String(30))
    reason_for_leaving_the_group = db.Column(db.String(30))
    bank_code = db.Column(db.String(9))
    bank_account_number = db.Column(db.String(10))
    swift_bic = db.Column(db.String(11))
    owner_of_bank_account  = db.Column(db.String(60))
    employee_with_potential  = db.Column(db.String(1))
    indicator_for_executive = db.Column(db.String(1))
    indicator_for_security_advisor = db.Column(db.String(1))
    date_entry = db.Column(db.Date,  default=date.today())
class AD_AN(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    indicator_eligibility_stock_purchase_program  = db.Column(db.String(1))
    company_code = db.Column(db.String(4))
    personal_number  = db.Column(db.String(9))
    plant_identifier_for_personal_number = db.Column(db.String(3))
    filler5 = db.Column(db.String(4))
    filler6 = db.Column(db.String(3))
    transfer_date = db.Column(db.String(30))
    filler7 = db.Column(db.String(4))
    filler8 = db.Column(db.String(3))
    local_cost_center = db.Column(db.String(8))
    date_entry = db.Column(db.Date,  default=date.today())
class AM_AW(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    employee_group = db.Column(db.String(1))
    filler9 = db.Column(db.String(30))
    department_abbreviation = db.Column(db.String(10))
    hr_executive_level = db.Column(db.String(3))
    employment_type = db.Column(db.String(3))
    organizational_code = db.Column(db.String(10))
    physical_work_location_code = db.Column(db.String(10))
    internal_mail_code = db.Column(db.String(10))
    level_of_business_allocation1 = db.Column(db.String(10))
    level_of_business_allocation2 = db.Column(db.String(10))
    date_entry = db.Column(db.Date,  default=date.today())
class AX_BG(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    mailing_language_of_employee = db.Column(db.String(2))
    building = db.Column(db.String(30))
    room_number = db.Column(db.String(10))
    location_code = db.Column(db.String(11))
    country_of_birth = db.Column(db.String(2))
    city_of_birth = db.Column(db.String(30))
    citizenship = db.Column(db.String(2))
    filler10 = db.Column(db.String(1))
    filler11 = db.Column(db.String(1))
    street_and_house_number_of_home_address = db.Column(db.String(40))
    date_entry = db.Column(db.Date,  default=date.today())

class BH_BQ(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    street_and_house_number_for_preferred_mailing_address = db.Column(db.String(2))
    filler12 = db.Column(db.String(30))
    job_code = db.Column(db.String(15))
    plant2 = db.Column(db.String(3))
    plant1 = db.Column(db.String(10))
    organizational_code2 = db.Column(db.String(10))
    smtp_email_address = db.Column(db.String(90))
    position_entry_date = db.Column(db.String(30))
    date_contract_end = db.Column(db.String(30))
    hr_department = db.Column(db.String(20))
    date_entry = db.Column(db.Date,  default=date.today())

class BR_CA(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    hr_representative = db.Column(db.String(30))
    fax_number_hr_representative = db.Column(db.String(25))
    management_level = db.Column(db.String(3))
    work_phone_number_of_hr_representative = db.Column(db.String(25))
    preferred_first_name = db.Column(db.String(30))
    middle_initial = db.Column(db.String(1))
    home_host_indicator = db.Column(db.String(1))
    fulltime_parttime_equivalency = db.Column(db.String(1))
    reports_to_kim= db.Column(db.String(15))
    date_of_entry_into_company = db.Column(db.String(10))
    date_entry = db.Column(db.Date,  default=date.today())

class CB_CW(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    filler12 = db.Column(db.String(3))
    filler13 = db.Column(db.String(3))
    filler14 = db.Column(db.String(3))
    filler15 = db.Column(db.String(3))
    filler16 = db.Column(db.String(3))
    filler17 = db.Column(db.String(3))
    filler18 = db.Column(db.String(3))
    filler19 = db.Column(db.String(3))
    filler20 = db.Column(db.String(3))
    filler21 = db.Column(db.String(3))
    filler22 = db.Column(db.String(3))
    filler23 = db.Column(db.String(3))
    filler24 = db.Column(db.String(3))
    filler25 = db.Column(db.String(3))
    filler26 = db.Column(db.String(3))
    filler27 = db.Column(db.String(3))
    filler28 = db.Column(db.String(3))
    filler29 = db.Column(db.String(3))
    filler30 = db.Column(db.String(3))
    filler31 = db.Column(db.String(3))
    filler32 = db.Column(db.String(3))
    filler33 = db.Column(db.String(3))
    date_entry = db.Column(db.Date,  default=date.today())
    
class CX_DG(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    current_position_title = db.Column(db.String(65))
    payment_type = db.Column(db.String(2))
    plant_section = db.Column(db.String(2))
    filler35 = db.Column(db.String(2))
    indicator_for_disabled_person = db.Column(db.String(1))
    social_security_number = db.Column(db.String(12))
    state = db.Column(db.String(3))
    state_for_preferred_mailing_address  = db.Column(db.String(3))
    phone_country_code = db.Column(db.String(4))    
    date_entry = db.Column(db.Date,  default=date.today())
    
class DH_DQ(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    phone_country_code = db.Column(db.String(4))    
    phone_extension_and_company = db.Column(db.String(12))
    fax_country_code = db.Column(db.String(4))
    fax_area_code = db.Column(db.String(10))
    fax_extension_and_company = db.Column(db.String(12))
    mobile_country_code = db.Column(db.String(4))    
    mobile_area_code = db.Column(db.String(10))
    mobile_number = db.Column(db.String(12))
    empl_id = db.Column(db.String(11))
    filler36 = db.Column(db.String(1))
    date_entry = db.Column(db.Date,  default=date.today())

class DR_EA(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    filler37 = db.Column(db.String(10))
    bv_wechsel_ab = db.Column(db.String(8))
    wohnhaft_bei = db.Column(db.String(40))
    company_code_expat = db.Column(db.String(4))
    currency = db.Column(db.String(3))
    filler38 = db.Column(db.String(10))
    year_of_goal_income = db.Column(db.String(4))
    monthly_salary = db.Column(db.String(10))
    delete_flag= db.Column(db.String(1))
    filler39 = db.Column(db.String(10))
    date_entry = db.Column(db.Date,  default=date.today())

class EB_EK(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    grade_band= db.Column(db.String(2))
    shift_indicator = db.Column(db.String(2))
    employee_category = db.Column(db.String(2))
    department_name = db.Column(db.String(60))
    status_employee = db.Column(db.String(1))
    code_dl_tv = db.Column(db.String(1))
    grade = db.Column(db.String(2))
    nacos_cost_center = db.Column(db.String(8))
    wage_group = db.Column(db.String(2))
    year_of_wage_group  = db.Column(db.String(4))
    date_entry = db.Column(db.Date,  default=date.today())
    
class EL_EY(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    date_of_assignment_to_wage_group = db.Column(db.String(8))
    filler40 = db.Column(db.String(10))
    logon_id_of_employee = db.Column(db.String(1))
    company_name_of_external_person = db.Column(db.String(38))
    filler41 = db.Column(db.String(4))
    filler42 = db.Column(db.String(8))    
    filler43 = db.Column(db.String(3))
    filler44 = db.Column(db.String(4))
    filler45 = db.Column(db.String(8))
    filler46 = db.Column(db.String(3))    
    filler47 = db.Column(db.String(4))
    filler48 = db.Column(db.String(8))
    filler49 = db.Column(db.String(3))
    filler50 = db.Column(db.String(3))
    date_entry = db.Column(db.Date,  default=date.today())
    
class EZ_FI(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    bv_wechsel_ab = db.Column(db.String(8))
    filler51 = db.Column(db.String(7))
    filler52 = db.Column(db.String(9))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))    
    date_entry = db.Column(db.Date,  default=date.today())
class FJ_FS(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    date_entry = db.Column(db.Date,  default=date.today())
class FT_GC(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    date_entry = db.Column(db.Date,  default=date.today())

class GD_GM(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    type_of_timeframe_for_tax_data = db.Column(db.String(1))
    starting_of_timeframe_for_tax_data = db.Column(db.String(8))
    end_of_timeframe_for_tax_data = db.Column(db.String(8))
    filler53 = db.Column(db.String(1))
    date_entry = db.Column(db.Date,  default=date.today())
class GN_GW(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    permanent_residence = db.Column(db.String(2))
    filler54 = db.Column(db.String(6))
    wage_and_salary_group = db.Column(db.String(3))
    diversity = db.Column(db.String(3))
    executive_bonus = db.Column(db.String(3))
    number_of_monthly_base_salary = db.Column(db.String(4))
    structure_code = db.Column(db.String(15))
    dept_id = db.Column(db.String(10))
    full_time_equivalent = db.Column(db.String(3))
    bonus_payout = db.Column(db.String(10))
    date_entry = db.Column(db.Date,  default=date.today())
class GX_HK(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    kim = db.Column(db.String(15),  nullable=False)
    position_control_number = db.Column(db.String(8))
    daimler_cost_center = db.Column(db.String(10))
    last_name_passport = db.Column(db.String(30))
    first_name_passport = db.Column(db.String(30))
    iban = db.Column(db.String(34))
    local_job_profile = db.Column(db.String(40))
    irwwh = db.Column(db.String(4))
    rabattkennzeichen = db.Column(db.String(3))
    valid_by = db.Column(db.String(8))
    buchungskreis = db.Column(db.String(4))
    widersprecherkennzeichen = db.Column(db.String(1))
    ivv_contract= db.Column(db.String(1))
    ivv_risk_taker = db.Column(db.String(1))
    ivv_control_unit = db.Column(db.String(1))
    date_entry = db.Column(db.Date,  default=date.today())
#---------------------------------------------------------------------------------------------------------------------
def load_data():
    file = 'static/uploads2/test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        db.session.add(GX_HK(kim=row[0].value,position_control_number=row[120].value,daimler_cost_center=row[120].value,last_name_passport=row[120].value,first_name_passport=row[120].value,iban=row[120].value,local_job_profile=row[120].value,irwwh=row[120].value,rabattkennzeichen=row[120].value,valid_by=row[120].value,buchungskreis=row[120].value,widersprecherkennzeichen=row[120].value,ivv_contract=row[120].value,ivv_risk_taker=row[120].value,ivv_control_unit=row[120].value))
        db.session.add(GN_GW(kim=row[0].value,permanent_residence=row[120].value,filler54=row[120].value,wage_and_salary_group=row[120].value,diversity=row[120].value,executive_bonus=row[120].value,number_of_monthly_base_salary=row[120].value,structure_code=row[120].value,dept_id=row[120].value,full_time_equivalent=row[203].value,bonus_payout=row[120].value))
        db.session.add(GD_GM(kim=row[0].value,starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value,filler53=row[120].value))
        db.session.add(FT_GC(kim=row[0].value,starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value))
        db.session.add(FJ_FS(kim=row[0].value,starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value))
        db.session.add(EZ_FI(kim=row[0].value,bv_wechsel_ab=row[120].value,filler51=row[120].value,filler52=row[120].value,type_of_timeframe_for_tax_data=row[120].value,starting_of_timeframe_for_tax_data=row[120].value,end_of_timeframe_for_tax_data=row[120].value))
        db.session.add(EL_EY(kim=row[0].value,date_of_assignment_to_wage_group=row[120].value,filler40=row[120].value,logon_id_of_employee=row[120].value,company_name_of_external_person=row[120].value,filler41=row[120].value,filler42=row[120].value,filler43=row[120].value,filler44=row[120].value,filler45=row[120].value,filler46=row[120].value,filler47=row[120].value,filler48=row[120].value,filler49=row[120].value,filler50=row[120].value))
        db.session.add(EB_EK(kim=row[0].value,grade_band=row[120].value,shift_indicator=row[120].value,employee_category=row[120].value,department_name=row[120].value,status_employee=row[120].value,code_dl_tv=row[120].value,grade=row[120].value,nacos_cost_center=row[120].value,wage_group=row[120].value,year_of_wage_group=row[120].value,))
        db.session.add(DR_EA(kim=row[0].value,filler37=row[120].value,bv_wechsel_ab=row[120].value,wohnhaft_bei=row[120].value,company_code_expat=row[120].value,currency=row[120].value,filler38=row[120].value,year_of_goal_income=row[120].value,monthly_salary=row[120].value,delete_flag=row[120].value,filler39=row[120].value))
        db.session.add(DH_DQ(kim=row[0].value,phone_country_code=row[111].value,phone_extension_and_company=row[112].value,fax_country_code=row[113].value,fax_area_code=row[114].value,fax_extension_and_company=row[115].value,mobile_country_code=row[116].value,mobile_area_code=row[117].value,mobile_number=row[118].value,empl_id=row[119].value,filler36=row[120].value))
        db.session.add(CX_DG(kim=row[0].value,current_position_title=row[101].value,payment_type=row[102].value,plant_section=row[103].value,filler35=row[104].value,indicator_for_disabled_person=row[105].value,social_security_number=row[106].value,state=row[108].value,state_for_preferred_mailing_address=row[109].value,phone_country_code=row[110].value))
        db.session.add(CB_CW(kim=row[0].value,filler12=row[79].value,filler13=row[80].value,filler14=row[81].value,filler15=row[82].value,filler16=row[83].value,filler17=row[84].value,filler18=row[85].value,filler19=row[86].value,filler20=row[87].value,filler21=row[88].value,filler22=row[89].value,filler23=row[90].value,filler24=row[91].value,filler25=row[92].value,filler26=row[93].value,filler27=row[94].value,filler28=row[95].value,filler29=row[96].value,filler30=row[97].value,filler31=row[98].value,filler32=row[99].value,filler33=row[100].value))
        db.session.add(BR_CA(kim=row[0].value,hr_representative=row[69].value,fax_number_hr_representative=row[70].value,management_level=row[71].value,work_phone_number_of_hr_representative=row[72].value,preferred_first_name=row[73].value,middle_initial=row[74].value,home_host_indicator=row[75].value,fulltime_parttime_equivalency =row[76].value,reports_to_kim=row[77].value,date_of_entry_into_company=row[78].value))
        db.session.add(BH_BQ(kim=row[0].value,street_and_house_number_for_preferred_mailing_address=row[59].value,filler12=row[60].value,job_code=row[61].value,plant1=row[63].value,plant2=row[62].value,organizational_code2=row[64].value,smtp_email_address=row[65].value,position_entry_date=row[66].value,date_contract_end=row[67].value,hr_department=row[68].value))
        db.session.add(AX_BG(kim=row[0].value,mailing_language_of_employee=row[49].value, building=row[50].value, room_number=row[51].value, location_code=row[52].value, country_of_birth=row[53].value, city_of_birth=row[54].value, citizenship=row[55].value, filler10=row[56].value,filler11=row[57].value, street_and_house_number_of_home_address=row[58].value))
        db.session.add(AM_AW(kim=row[0].value,employee_group=row[39].value,filler9=row[40].value,department_abbreviation=row[41].value,hr_executive_level=row[42].value,employment_type=row[43].value,organizational_code=row[44].value,physical_work_location_code=row[44].value,internal_mail_code=row[45].value,level_of_business_allocation1=row[46].value,level_of_business_allocation2=row[47].value))
        db.session.add(AD_AN(kim=row[0].value,indicator_eligibility_stock_purchase_program=row[29].value, company_code=row[30].value, personal_number=row[32].value,plant_identifier_for_personal_number=row[31].value,filler5=row[33].value,filler6=row[34].value,transfer_date=row[35].value,filler7=row[36].value,filler8=row[37].value,local_cost_center=row[38].value))
        db.session.add(A_I(kim=row[0].value, sin=row[1].value,gender=row[2].value,first_name=row[3].value,last_name=row[4].value,last_name_prefix=row[5].value,academic_title=row[6].value,filler1=row[7].value,birth_date=row[8].value))
        db.session.add(J_S(kim=row[0].value,country_code=row[9].value,postal_code=row[10].value,city=row[11].value,c_o=row[12].value,filler2=row[13].value,filler3=row[14].value,filler4=row[15].value,country_code_mail_address=row[16].value,postal_code_mail_address=row[17].value,city_mail_address=row[18].value))
        db.session.add(T_AC(kim=row[0].value,date_of_entry_into_group=row[19].value,date_of_leaving_group=row[20].value,reason_for_leaving_the_group=row[21].value,bank_code=row[22].value,bank_account_number=row[23].value,swift_bic=row[24].value,owner_of_bank_account=row[25].value,employee_with_potential=row[26].value,indicator_for_executive=row[27].value,indicator_for_security_advisor=row[28].value))
        print('loading...')
        
        db.session.commit()
#-----------------------------------------------------------------------------------------------------------------------------------
def save_emp(data):
        
    file = 'static/uploads2/test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=10, min_col=1, max_row=11, max_col=ws.max_column):
        db.session.add(GX_HK(kim=data.get('kim'),position_control_number=row[120].value,daimler_cost_center=row[120].value,last_name_passport=row[120].value,first_name_passport=row[120].value,iban=row[120].value,local_job_profile=row[120].value,irwwh=row[120].value,rabattkennzeichen=row[120].value,valid_by=row[120].value,buchungskreis=row[120].value,widersprecherkennzeichen=row[120].value,ivv_contract=row[120].value,ivv_risk_taker=row[120].value,ivv_control_unit=row[120].value))
        db.session.add(GN_GW(kim=data.get('kim'),permanent_residence=row[120].value,filler54=row[120].value,wage_and_salary_group=row[120].value,diversity=row[120].value,executive_bonus=row[120].value,number_of_monthly_base_salary=row[120].value,structure_code=row[120].value,dept_id=row[120].value,full_time_equivalent=row[203].value,bonus_payout=row[120].value))
        db.session.add(GD_GM(kim=data.get('kim'),starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value,filler53=row[120].value))
        db.session.add(FT_GC(kim=data.get('kim'),starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value))
        db.session.add(FJ_FS(kim=data.get('kim'),starting_of_timeframe_for_tax_data=row[120].value, end_of_timeframe_for_tax_data=row[120].value,type_of_timeframe_for_tax_data=row[120].value))
        db.session.add(EZ_FI(kim=data.get('kim'),bv_wechsel_ab=row[120].value,filler51=row[120].value,filler52=row[120].value,type_of_timeframe_for_tax_data=row[120].value,starting_of_timeframe_for_tax_data=row[120].value,end_of_timeframe_for_tax_data=row[120].value))
        db.session.add(EL_EY(kim=data.get('kim'),date_of_assignment_to_wage_group=row[120].value,filler40=row[120].value,logon_id_of_employee=row[120].value,company_name_of_external_person=row[120].value,filler41=row[120].value,filler42=row[120].value,filler43=row[120].value,filler44=row[120].value,filler45=row[120].value,filler46=row[120].value,filler47=row[120].value,filler48=row[120].value,filler49=row[120].value,filler50=row[120].value))
        db.session.add(EB_EK(kim=data.get('kim'),grade_band=row[120].value,shift_indicator=row[120].value,employee_category=row[120].value,department_name=row[120].value,status_employee=row[120].value,code_dl_tv=row[120].value,grade=row[120].value,nacos_cost_center=row[120].value,wage_group=row[120].value,year_of_wage_group=row[120].value,))
        db.session.add(DR_EA(kim=data.get('kim'),filler37=row[120].value,bv_wechsel_ab=row[120].value,wohnhaft_bei=row[120].value,company_code_expat=row[120].value,currency=row[120].value,filler38=row[120].value,year_of_goal_income=row[120].value,monthly_salary=row[120].value,delete_flag=row[120].value,filler39=row[120].value))
        db.session.add(DH_DQ(kim=data.get('kim'),phone_country_code=row[111].value,phone_extension_and_company=row[112].value,fax_country_code=row[113].value,fax_area_code=row[114].value,fax_extension_and_company=row[115].value,mobile_country_code=row[116].value,mobile_area_code=row[117].value,mobile_number=row[118].value,empl_id=row[119].value,filler36=row[120].value))
        db.session.add(CX_DG(kim=data.get('kim'),current_position_title=data.get('position'),payment_type=row[102].value,plant_section=row[103].value,filler35=row[104].value,indicator_for_disabled_person=row[105].value,social_security_number=row[106].value,state=row[108].value,state_for_preferred_mailing_address=row[109].value,phone_country_code=row[110].value))
        db.session.add(CB_CW(kim=data.get('kim'),filler12=row[79].value,filler13=row[80].value,filler14=row[81].value,filler15=row[82].value,filler16=row[83].value,filler17=row[84].value,filler18=row[85].value,filler19=row[86].value,filler20=row[87].value,filler21=row[88].value,filler22=row[89].value,filler23=row[90].value,filler24=row[91].value,filler25=row[92].value,filler26=row[93].value,filler27=row[94].value,filler28=row[95].value,filler29=row[96].value,filler30=row[97].value,filler31=row[98].value,filler32=row[99].value,filler33=row[100].value))
        db.session.add(BR_CA(kim=data.get('kim'),hr_representative=row[69].value,fax_number_hr_representative=row[70].value,management_level=row[71].value,work_phone_number_of_hr_representative=row[72].value,preferred_first_name=row[73].value,middle_initial=row[74].value,home_host_indicator=row[75].value,fulltime_parttime_equivalency =row[76].value,reports_to_kim=data.get('report'),date_of_entry_into_company=row[78].value))
        db.session.add(BH_BQ(kim=data.get('kim'),street_and_house_number_for_preferred_mailing_address=row[59].value,filler12=row[60].value,job_code=row[61].value,plant1=row[63].value,plant2=row[62].value,organizational_code2=row[64].value,smtp_email_address=data.get('temail'),position_entry_date=row[66].value,date_contract_end=row[67].value,hr_department=row[68].value))
        db.session.add(AX_BG(kim=data.get('kim'),mailing_language_of_employee=row[49].value, building=row[50].value, room_number=row[51].value, location_code=row[52].value, country_of_birth=row[53].value, city_of_birth=row[54].value, citizenship=row[55].value, filler10=row[56].value,filler11=row[57].value, street_and_house_number_of_home_address=row[58].value))
        db.session.add(AM_AW(kim=data.get('kim'),employee_group=data.get('egroup'),filler9=row[40].value,department_abbreviation=row[41].value,hr_executive_level=data.get('hrlevel'),employment_type=data.get('etype'),organizational_code=row[44].value,physical_work_location_code=row[44].value,internal_mail_code=row[45].value,level_of_business_allocation1=row[46].value,level_of_business_allocation2=row[47].value))
        db.session.add(AD_AN(kim=data.get('kim'),indicator_eligibility_stock_purchase_program=row[29].value, company_code=row[30].value, personal_number=data.get('pnumber'),plant_identifier_for_personal_number=row[31].value,filler5=row[33].value,filler6=row[34].value,transfer_date=row[35].value,filler7=row[36].value,filler8=row[37].value,local_cost_center=row[38].value))
        db.session.add(T_AC(kim=data.get('kim'),date_of_entry_into_group=data.get('entry'),date_of_leaving_group=row[20].value,reason_for_leaving_the_group=row[21].value,bank_code=row[22].value,bank_account_number=row[23].value,swift_bic=row[24].value,owner_of_bank_account=row[25].value,employee_with_potential=row[26].value,indicator_for_executive=row[27].value,indicator_for_security_advisor=row[28].value))
        db.session.add(J_S(kim=data.get('kim'),country_code=row[9].value,postal_code=row[10].value,city=row[11].value,c_o=row[12].value,filler2=row[13].value,filler3=row[14].value,filler4=row[15].value,country_code_mail_address=row[16].value,postal_code_mail_address=row[17].value,city_mail_address=row[18].value))
        db.session.add(A_I(kim=data.get('kim'), sin='MDAT589   ',gender=data.get('gender'),first_name=data.get('fname'),last_name=data.get('lname'),last_name_prefix='                              ',academic_title='                              ',filler1='                              ',birth_date=data.get('bdate')))
        db.session.commit()
        break
#---------------------------------------------------------------------------------------------------------------------
def process(wording, spacing):
    max_len = len(wording)
    if max_len > spacing:
        return wording[0:spacing]
    padding = spacing-max_len
    for i in range(0, padding):
        wording = wording+" "
    return wording
#---------------------------------------------------------------------------------------------------------------------

def fix(title,spacing):
    X = 'New Kim	MDAT	Gender	Last Name	First Name	Academic Title	Birth Date	Date of Entry into Group	Date of Leaving Group	Reason for Leaving the Group	Indicator for Executive	Company Code	Plant identifier for Personal Number	Personal Number	Transfer Date	Cost Center	Employee group	Department Abbreviation	HR Executive Level	Employment Type	STELLENIDENT	Physical work location code	Level of business allocation 1	Level of business allocation 2	Physical work location code	GEBURTSLAND	STAATSANGEH	Currency	MONATSGEHALT	Plant 2/ Center	Plant 1	Position Number	Position Entry Date	HR Department	HR Representative	Management Level	Middle Initial	Home/Host Indicator	Fulltime/ Parttime equivalency	BERICHTETAN	Date of Entry into Company	KZBEHINDERT	Empl ID	NULL	GI_EXIMPAT	ENTGELTGRUPPE	EXECUTIVE_BONUS	    Diversity	ANZ_MONATSGEH	Strukturkennzahl	Dept ID	FTE	Confirmation Status	Plant section	Job Code	Payment Type	SCHICHT	BESCHGR	STATUS_ABR	VERSANDART	RANGSTUFE	NACOS_KOST	TG	TG_BJAHR	GUEAB	SUCHNACHNAME	SUCHVORNAME	BONUS_PAYOUT	Assigned D'
    c = X.replace("	",',')
    c =c.split(',')
    b=[]
    cx = ''
    for v in c:
            max_len = spacing-len(v)
            b.append(v+ max_len*' ')
    for e in b:
        cx = cx+e
    with open('static/reports2/'+title+'.txt','w', encoding = 'utf-8') as f:
            f.write(cx)
#---------------------------------------------------------------------------------------------------------------------
def gen_report2(obj,spacing,title):
    data= []
    spacing = 12
    a=obj.get('A_I')
    j = obj.get('J_S')
    t =obj.get('T_AC')
    ad =obj.get('AD_AN')
    am =obj.get('AM_AW')
    ax =obj.get('AX_BG')
    bh =obj.get('BH_BQ')
    br =obj.get('BR_CA')
    cb =obj.get('CB_CW')
    cx =obj.get('CX_DG')
    dh =obj.get('DH_DQ')
    dr=obj.get('DR_EA')
    eb =obj.get('EB_EK')
    el = obj.get('EL_EY')
    ez =obj.get('EZ_FI')
    fj =obj.get('FJ_FS')
    ft =obj.get('FT_GC')
    gd =obj.get('GD_GM')
    gn =obj.get('GN_GW')
    gx =obj.get('GX_HK')
    #fix(title,int(spacing))
    for c in range(0,len(obj.get('A_I'))-1):
        data.append([a[c].kim, a[c].sin,a[c].gender,a[c].first_name,a[c].last_name,a[c].last_name_prefix,a[c].academic_title,a[c].filler1,a[c].birth_date,
                                j[c].country_code, j[c].postal_code,j[c].city,j[c].c_o,j[c].filler2,j[c].filler3,j[c].filler4,j[c].country_code_mail_address,j[c].postal_code_mail_address,j[c].city_mail_address,
                                t[c].date_of_entry_into_group,t[c].date_of_leaving_group,t[c].reason_for_leaving_the_group,t[c].bank_code,t[c].bank_account_number,t[c].swift_bic, t[c].owner_of_bank_account,t[c].employee_with_potential,t[c].indicator_for_executive,t[c].indicator_for_security_advisor,
                                ad[c].indicator_eligibility_stock_purchase_program,ad[c].company_code,ad[c].personal_number,ad[c].plant_identifier_for_personal_number,ad[c].filler5,ad[c].filler6,ad[c].filler5,ad[c].transfer_date,ad[c].filler7,ad[c].filler8,ad[c].local_cost_center,
                                am[c].employee_group,am[c].filler9,am[c].department_abbreviation,am[c].hr_executive_level,am[c].employment_type, am[c].organizational_code,am[c].physical_work_location_code, am[c].internal_mail_code, am[c].level_of_business_allocation1,am[c].level_of_business_allocation2,
                                ax[c].mailing_language_of_employee, ax[c].building,ax[c].room_number,ax[c].country_of_birth,ax[c].city_of_birth,ax[c].citizenship,ax[c].filler10,ax[c].filler11,ax[c].street_and_house_number_of_home_address,
                                bh[c].street_and_house_number_for_preferred_mailing_address,bh[c].filler12,bh[c].job_code,bh[c].plant2,bh[c].plant1,bh[c].organizational_code2,bh[c].smtp_email_address,bh[c].position_entry_date,bh[c].date_contract_end,bh[c].hr_department,
                                br[c].hr_representative,br[c].fax_number_hr_representative,br[c].management_level,br[c].work_phone_number_of_hr_representative,br[c].preferred_first_name,br[c].middle_initial,br[c].home_host_indicator,br[c].fulltime_parttime_equivalency,br[c].reports_to_kim,br[c].date_of_entry_into_company,
                                cb[c].filler12,cb[c].filler13,cb[c].filler14,cb[c].filler15,cb[c].filler16,cb[c].filler17,cb[c].filler18,cb[c].filler19,cb[c].filler20,cb[c].filler21,cb[c].filler21,cb[c].filler22,cb[c].filler23,cb[c].filler24,cb[c].filler25,cb[c].filler26,cb[c].filler26,cb[c].filler27,cb[c].filler28,cb[c].filler29,cb[c].filler30,cb[c].filler31,cb[c].filler32,cb[c].filler33,
                                cx[c].current_position_title,cx[c].payment_type,cx[c].plant_section,cx[c].filler35,cx[c].indicator_for_disabled_person,cx[c].social_security_number,cx[c].state,cx[c].state_for_preferred_mailing_address,cx[c].phone_country_code,
                                dh[c].phone_country_code,dh[c].phone_extension_and_company,dh[c].fax_country_code,dh[c].fax_area_code,dh[c].fax_extension_and_company,dh[c].mobile_country_code,dh[c].mobile_area_code,dh[c].mobile_number,dh[c].empl_id,dh[c].filler36,
                                dr[c].filler37,dr[c].bv_wechsel_ab,dr[c].wohnhaft_bei,dr[c].company_code_expat,dr[c].currency,dr[c].filler38,dr[c].year_of_goal_income,dr[c].monthly_salary,dr[c].delete_flag,dr[c].filler39,
                                eb[c].grade_band,eb[c].shift_indicator,eb[c].employee_category,eb[c].department_name,eb[c].status_employee,eb[c].code_dl_tv,eb[c].grade,eb[c].nacos_cost_center,eb[c].wage_group,eb[c].year_of_wage_group,
                                el[c].date_of_assignment_to_wage_group,el[c].filler40,el[c].logon_id_of_employee,el[c].company_name_of_external_person,el[c].filler41,el[c].filler42,el[c].filler43,el[c].filler44,el[c].filler45,el[c].filler46,el[c].filler47,el[c].filler48,el[c].filler49,el[c].filler50,
                                ez[c].bv_wechsel_ab,ez[c].filler51,ez[c].filler52,ez[c].type_of_timeframe_for_tax_data,ez[c].starting_of_timeframe_for_tax_data,ez[c].end_of_timeframe_for_tax_data,ez[c].type_of_timeframe_for_tax_data,ez[c].starting_of_timeframe_for_tax_data,ez[c].end_of_timeframe_for_tax_data,ez[c].type_of_timeframe_for_tax_data,
                                fj[c].starting_of_timeframe_for_tax_data,fj[c].end_of_timeframe_for_tax_data,fj[c].type_of_timeframe_for_tax_data,fj[c].starting_of_timeframe_for_tax_data,fj[c].end_of_timeframe_for_tax_data,fj[c].type_of_timeframe_for_tax_data,fj[c].starting_of_timeframe_for_tax_data,fj[c].end_of_timeframe_for_tax_data,fj[c].type_of_timeframe_for_tax_data,fj[c].starting_of_timeframe_for_tax_data,
                                ft[c].end_of_timeframe_for_tax_data,ft[c].type_of_timeframe_for_tax_data,ft[c].starting_of_timeframe_for_tax_data,ft[c].end_of_timeframe_for_tax_data,ft[c].type_of_timeframe_for_tax_data,ft[c].starting_of_timeframe_for_tax_data,ft[c].end_of_timeframe_for_tax_data,ft[c].type_of_timeframe_for_tax_data,ft[c].starting_of_timeframe_for_tax_data,ft[c].end_of_timeframe_for_tax_data,
                                gd[c].type_of_timeframe_for_tax_data,gd[c].starting_of_timeframe_for_tax_data,gd[c].end_of_timeframe_for_tax_data,gd[c].type_of_timeframe_for_tax_data,gd[c].starting_of_timeframe_for_tax_data,gd[c].end_of_timeframe_for_tax_data,gd[c].type_of_timeframe_for_tax_data,gd[c].starting_of_timeframe_for_tax_data,gd[c].end_of_timeframe_for_tax_data,gd[c].filler53,
                                gn[c].permanent_residence,gn[c].filler54,gn[c].wage_and_salary_group,gn[c].diversity,gn[c].executive_bonus,gn[c].number_of_monthly_base_salary,gn[c].structure_code,gn[c].dept_id,gn[c].full_time_equivalent,gn[c].bonus_payout,
                                gx[c].position_control_number,gx[c].daimler_cost_center,gx[c].last_name_passport,gx[c].first_name_passport,gx[c].iban,gx[c].local_job_profile,gx[c].irwwh,gx[c].rabattkennzeichen,gx[c].valid_by,gx[c].buchungskreis,gx[c].widersprecherkennzeichen,gx[c].ivv_contract,gx[c].ivv_risk_taker,gx[c].ivv_control_unit,
                                ])
    with open('static/reports2/'+title+'.txt','a', encoding = 'utf-8') as f:
        for i in data:
            f.write('\n')
            for j in i:
                f.writelines(process(str(j),spacing).replace('.',''))
#---------------------------------------------------------------------------------------------------------------------
def test_gen(obj,spacing,title):
    data= []
    spacing = 16
    fix(title,spacing)
    with open('static/reports2/'+title+'.txt','a', encoding = 'utf-8') as f:
        for item in obj:
            for xy in item:
                f.write('\n')
                for attr,value in vars(xy).items():
                    if  attr.startswith('_')==False  and attr!='id' and attr!='kim' and attr!='date_entry' :
                        #data.append(value)
                        f.writelines(process(str(value),spacing).replace('.',''))
                        print(attr)
                
        
        #for i in data:
        #    f.write('\n')
       #     print(i)
      #      f.writelines(process(str(i),spacing).replace('.',''))

     
#---------------------------------------------------------------------------------------------------------------------

def save_profile(data):
    try:
        with open('static/profiles/profiles.json','w', encoding = 'utf-8') as f:
            json.dump(data, f)
            f.close()
            return 1
    except Exception as e:
        print(e)
        return 0
    
#---------------------------------------------------------------------------------------------------------------------
def get_profile():
    data = {}
    with open('static/profiles/profiles.json','r',encoding = 'utf-8') as f:
        try:
            data = json.load(f)        
        except Exception as e:
            print(e)
    return data
#---------------------------------------------------------------------------------------------------------------------
def get_data(path):
    wb = load_workbook(path)
    ws = wb.active
    AidsUSD =0
    AidsLevy = 0
    TaxUSD = 0
    GrossIncomeUSD = 0
    totalFringeB = 0
    AidsZWL =0
    PAYETax= 0
    PAYETaxZWL=0
    totalEarningsUSD=0
    totalEarningsZWL = 0
    cols = [6,8,9,10,11,12,13,14,15,16]
    for i in cols:
        for x in (range(6,ws.max_row+1)):
            if i == 6:
                try:
                    AidsUSD = AidsUSD + ws.cell(row=x, column=i).value
                    
                    
                except:
                        continue
            if i == 8:
                try:
                    GrossIncomeUSD = GrossIncomeUSD + ws.cell(row=x, column=i).value
                    
                except:
                        continue
            if i == 9:
                try:
                    TaxUSD = TaxUSD + ws.cell(row=x, column=i).value
                    
                except:
                        continue
            if i == 10:
                try:
                    AidsZWL = AidsZWL + ws.cell(row=x, column=i).value
                except:
                        continue

            if i == 11:
                try:
                    PAYETax = PAYETax + ws.cell(row=x, column=i).value
                except:
                        continue
            if i == 12:
                try:
                    totalEarningsUSD = totalEarningsUSD + ws.cell(row=x, column=i).value
                except:
                        continue
            if i == 13:
                try:
                    AidsLevy = AidsLevy + ws.cell(row=x, column=i).value
                except:
                        continue
            if i == 14:
                try:
                    PAYETaxZWL = PAYETaxZWL + ws.cell(row=x, column=i).value
                except:
                        continue

            if i == 15:
                try:
                    totalEarningsZWL = totalEarningsZWL + ws.cell(row=x, column=i).value
                except:
                        continue

            if i == 16:
                try:
                    totalFringeB = totalFringeB + ws.cell(row=x, column=i).value
                except:
                        continue

#--------------------------------------------------
    values = {}
    values['tr1'] = '{:.2f}'.format((totalEarningsUSD+totalEarningsZWL+totalFringeB))
    values['tr2'] =  '{:.2f}'.format((totalEarningsZWL+totalFringeB))
    values['tr3'] = '{:.2f}'.format((totalEarningsUSD))
    values['tr4'] = '{:.2f}'.format((GrossIncomeUSD))
    values['ne'] = ws.max_row-6
    values['gp1'] = '{:.2f}'.format((PAYETax+ PAYETaxZWL))
    values['gp2'] = '{:.2f}'.format((PAYETaxZWL))
    values['gp3'] = '{:.2f}'.format((PAYETax))
    values['gp4'] = '{:.2f}'.format((TaxUSD))
    values['al1'] = '{:.2f}'.format((AidsZWL+AidsLevy))
    values['al2'] = '{:.2f}'.format((AidsLevy))
    values['al3'] = '{:.2f}'.format((AidsZWL))
    values['al4'] = '{:.2f}'.format((AidsUSD))
    values['tt1'] = '{:.2f}'.format((PAYETax+PAYETaxZWL+AidsZWL+AidsLevy))
    values['tt2'] = '{:.2f}'.format((PAYETaxZWL+AidsLevy))
    values['tt3'] = '{:.2f}'.format((PAYETax+AidsZWL))
    values['tt4'] = '{:.2f}'.format((TaxUSD+AidsUSD))
    return values

def gen_report(data3, data2):
    data1 = get_profile()
    document = MailMerge('static/templates/temp.docx')
    document.merge(
            ename = data1.get('ename'),
            tname = data1.get('tname'),
            btnum = data1.get('btnum'),
            paye = data1.get('paye'),
            tin = data1.get('tin'),
            address = data1.get('address'),
            postal=data1.get('postal'),
            cell= data1.get('cell'),
            email= data1.get('email'),
            tax_period= data3.get('tax_period'),
            due_date= data3.get('due_date'),
            rate= data3.get('rate'),
            region= data1.get('region'),
            station= data1.get('station'),
            tr1 = ' '+str(data2.get('tr1')),
            tr2 = str(data2.get('tr2')),
            tr3 = str(data2.get('tr3')),
            tr4 = str(data2.get('tr4')),
            ne = str(data1.get('num')),
            gp1 = str(data2.get('gp1')),
            gp2 = str(data2.get('gp2')),
            gp3 = str(data2.get('gp3')),
            gp4 = str(data2.get('gp4')),
            al1 = str(data2.get('al1')),
            al2 = str(data2.get('al2')),
            al3 = str(data2.get('al3')),
            al4 = str(data2.get('al4')),
            tt1 = str(data2.get('tt1')),
            tt2 = str(data2.get('tt2')),
            tt3 = str(data2.get('tt3')),
            tt4 = str(data2.get('tt4')),
        )
    title = data3.get('title')
  
    doc_path = 'static/reports/'+str(title)+str('.docx')
    try:
        os.remove(os.path.join(doc_path))
        document.write(doc_path)
    except:
        document.write(doc_path)
  
#-------------------------------------------------------------
@app.route('/', methods=['POST','GET'])
def index():
   
    error = 'none'
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        if email =='client@petalmafrica.com' and password == 'client@123':
            return home('none')
        else:
            error = 'invalid log in details'
    template = 'sign-in.html'
    return render_template(template, error = error)

@app.route('/new_report/', methods=['POST','GET'])
def new_report():
    if request.files and request.method == 'POST' :
        file = request.files['file']
        try:
            os.remove(os.path.join("static/uploads/", file.filename))
            file.save(os.path.join("static/uploads/", file.filename))
        except:
            file.save(os.path.join("static/uploads/", file.filename))
        data = request.form.to_dict()
        try:
            #print(get_data(os.path.join("static/uploads/", file.filename)))
            gen_report(data, get_data(os.path.join("static/uploads/", file.filename)))
            return reports('Report was saved successfully')
        except Exception as e:
            print(e)
            return home('failed to save Report')
    template = 'index.html'
    return render_template(template)

@app.route('/home/<string:error>/')
def home(error):
    template = 'home.html'
    return render_template(template,error=error)

@app.route('/update_profile/', methods=['POST','GET'])
def update_profile():
    if request.method == 'POST':
        data = request.form.to_dict()
        save_profile(data)
        return home('Profile Update Successful')
    template = 'profiles.html'
    data = get_profile()
    return render_template(template,data=data)

@app.route('/add_new_employee/', methods=['POST','GET'])
def add_new_employee():
    if request.method == 'POST':
        data = request.form.to_dict()
        save_emp(data)
        return home('Employee Information Successfully')
    template = 'manage.html'
    data = get_profile()
    return render_template(template,data=data)

@app.route('/get_emp/', methods=['POST','GET'])
def get_emp():
    error=None
    try:
        if request.method == 'GET':
            data = request.args.to_dict()
            emp = A_I.query.filter_by(kim=data.get('kim')).first()
            if emp!=None and data.get("kim")!=None:
                error =None
                return update_data(data.get("kim"))
            elif emp==None and data.get("kim")!=None:
                        error ='Could not find employee "'+str(data.get('kim'))+'" '
            else:
                    error =None
        template = 'search.html'
        return render_template(template,error=error)
    except Exception as e:
        print(e)
        return home(e)


@app.route('/update_data<string:emp>/', methods=['POST','GET'])
def update_data(emp):
    emp1 = A_I.query.filter_by(kim=emp).first()
    emp2 = T_AC.query.filter_by(kim=emp).first()
    emp3 = AD_AN.query.filter_by(kim=emp).first()
    emp4 = AM_AW.query.filter_by(kim=emp).first()
    emp5 = BH_BQ.query.filter_by(kim=emp).first()
    emp6 = BR_CA.query.filter_by(kim=emp).first()
    emp7 = CX_DG.query.filter_by(kim=emp).first()
    emp8 = DH_DQ.query.filter_by(kim=emp).first()
    emp9 = DR_EA.query.filter_by(kim=emp).first()
    emp10 = EB_EK.query.filter_by(kim=emp).first()
    emp11 = EL_EY.query.filter_by(kim=emp).first()
    emp12 = EZ_FI.query.filter_by(kim=emp).first()
    emp13 = FJ_FS.query.filter_by(kim=emp).first()
    emp14 = FT_GC.query.filter_by(kim=emp).first()
    emp15 = GD_GM.query.filter_by(kim=emp).first()
    emp16 = GN_GW.query.filter_by(kim=emp).first()
    emp17 = GX_HK.query.filter_by(kim=emp).first()
    emp18= J_S.query.filter_by(kim=emp).first()
    emp19 = AX_BG.query.filter_by(kim=emp).first()
    
    if request.method == 'POST':
        data = request.form.to_dict()
        emp1.kim = data.get("kim")
        emp1.first_name= data.get("fname")
        emp1.last_name= data.get("lname")
        emp1.birth_date= data.get("bdate")
        emp2.date_of_entry_into_group= data.get("entry")
        emp3.personal_number= data.get("pnumber")
        emp4.employee_group= data.get("egroup")
        emp4.hr_executive_level= data.get("hrlevel")
        emp4.employment_type= data.get("etype")
        emp5.smtp_email_address= data.get("temail")
        emp6.reports_to_kim= data.get("report")
        emp7.current_position_title= data.get("position")
        emp1.date_entry = date.today()
        emp2.date_entry = date.today()
        emp3.date_entry = date.today()
        emp4.date_entry = date.today()
        emp5.date_entry = date.today()
        emp6.date_entry = date.today()
        emp7.date_entry = date.today()
        emp8.date_entry = date.today()
        emp9.date_entry = date.today()
        emp10.date_entry = date.today()
        emp11.date_entry = date.today()
        emp12.date_entry = date.today()
        emp13.date_entry = date.today()
        emp14.date_entry = date.today()
        emp15.date_entry = date.today()
        emp16.date_entry = date.today()
        emp17.date_entry = date.today()
        emp18.date_entry = date.today()
        emp19.date_entry = date.today()
        db.session.commit()
        return home('Information Update Successful')
    template = 'update.html'
    emp_data = {'kim':emp1.kim,'fname':emp1.first_name,'lname':emp1.last_name,'bdate':emp1.birth_date,'entry':emp2.date_of_entry_into_group,'pnumber':emp3.personal_number,'egroup':emp4.employee_group,'hrlevel':emp4.hr_executive_level,'etype':emp4.employment_type,'stemail':emp5.smtp_email_address,'report':emp6.reports_to_kim,'position':emp7.current_position_title}

    return render_template(template,data=emp_data)

@app.route("/reports/<string:error>/")
def reports(error):
    try:
        onlyfiles = [f for f in os.listdir('static/reports')]
        template = 'reports.html'
        return render_template(template,files=onlyfiles, error = error)

    except Exception as e:
        print(e)    
        template = 'reports.html'
        return home(e)


@app.route("/download/<string:name>")
def download(name):
    download_file = name
    print('downloading ',name)
    try:
        return send_from_directory(directory='static/reports',filename=download_file) 
    except Exception as e:
        print(e)
        return reports('failed to download file')


@app.route("/delete/<string:name>")
def delete(name):
    error = 'Document Deleted Successfully'
    try:
        os.remove('static/reports/'+name)
    except Exception as e:
        print(e)
        error = 'Failed to delete File'
    print('deleting ',name)
    return reports(error)

@app.route('/new_report2/', methods=['POST','GET'])
def new_report2():
    try:
        if  request.method == 'POST' :
            data = request.form.to_dict()
            start = data.get('start')
            stop = data.get('stop')
            rep = A_I.query.filter(A_I.date_entry > func.DATE(start), A_I.date_entry < func.DATE(stop)).all()
            rep1 = J_S.query.filter(J_S.date_entry > func.DATE(start), J_S.date_entry < func.DATE(stop)).all()
            rep2 = T_AC.query.filter(T_AC.date_entry > func.DATE(start), T_AC.date_entry < func.DATE(stop)).all()
            rep3 = AD_AN.query.filter(AD_AN.date_entry > func.DATE(start),AD_AN.date_entry < func.DATE(stop)).all()
            rep4 = AM_AW.query.filter(AM_AW.date_entry > func.DATE(start), AM_AW.date_entry < func.DATE(stop)).all()
            rep5 = AX_BG.query.filter(AX_BG.date_entry > func.DATE(start), AX_BG.date_entry < func.DATE(stop)).all()
            rep6 = BH_BQ.query.filter(BH_BQ.date_entry > func.DATE(start), BH_BQ.date_entry < func.DATE(stop)).all()
            rep7 = BR_CA.query.filter(BR_CA.date_entry > func.DATE(start), BR_CA.date_entry < func.DATE(stop)).all()
            rep8 = CB_CW.query.filter(CB_CW.date_entry > func.DATE(start), CB_CW.date_entry < func.DATE(stop)).all()
            rep9 = CX_DG.query.filter(CX_DG.date_entry > func.DATE(start), CX_DG.date_entry < func.DATE(stop)).all()
            rep10 = DH_DQ.query.filter(DH_DQ.date_entry > func.DATE(start), DH_DQ.date_entry < func.DATE(stop)).all()
            rep11 = DR_EA.query.filter(DR_EA.date_entry > func.DATE(start), DR_EA.date_entry < func.DATE(stop)).all()
            rep12 = EB_EK.query.filter(EB_EK.date_entry > func.DATE(start), EB_EK.date_entry < func.DATE(stop)).all()
            rep13 = EL_EY.query.filter(EL_EY.date_entry > func.DATE(start), EL_EY.date_entry < func.DATE(stop)).all()
            rep14 = EZ_FI.query.filter(EZ_FI.date_entry > func.DATE(start), EZ_FI.date_entry < func.DATE(stop)).all()
            rep15 = FJ_FS.query.filter(FJ_FS.date_entry > func.DATE(start), FJ_FS.date_entry < func.DATE(stop)).all()
            rep16 = FT_GC.query.filter(FT_GC.date_entry > func.DATE(start), FT_GC.date_entry < func.DATE(stop)).all()
            rep17 = GD_GM.query.filter(GD_GM.date_entry > func.DATE(start), GD_GM.date_entry < func.DATE(stop)).all()
            rep18 = GN_GW.query.filter(GN_GW.date_entry > func.DATE(start), GN_GW.date_entry < func.DATE(stop)).all()
            rep19 = GX_HK.query.filter(GX_HK.date_entry > func.DATE(start), GX_HK.date_entry < func.DATE(stop)).all()
            print('rows: ',len(rep))
            obj= {}
            obj['A_I'] =rep
            obj['J_S'] = rep1
            obj['T_AC'] = rep2
            obj['AD_AN'] = rep3
            obj['AM_AW'] = rep4
            obj['AX_BG'] = rep5
            obj['BH_BQ'] = rep6
            obj['BR_CA'] = rep7
            obj['CB_CW'] = rep8
            obj['CX_DG'] = rep9
            obj['DH_DQ'] = rep10
            obj['DR_EA'] = rep11
            obj['EB_EK'] = rep12
            obj['EL_EY'] = rep13
            obj['EZ_FI'] = rep14
            obj['FJ_FS'] = rep15
            obj['FT_GC'] = rep16
            obj['GD_GM'] = rep17
            obj['GN_GW'] = rep18
            obj['GX_HK'] = rep19     
            if len(rep)<1:
                return reports2('No data was saved within this time period')
            else:
                try:
                    gen_report2(obj,data.get('spacing'),data.get('title'))
            
                    return reports2('Report was saved successfully')
                except Exception as e:
                    print(e)
                return home('failed to save Report, please try again with different dates')
        template = 'index2.html'
        return render_template(template)
    except Exception as e:
        print(e)
        return home(e)


@app.route("/download2/<string:name>")
def download2(name):
        download_file = name
        print('downloading ',name)
        try:
            return send_from_directory(directory='static/reports2',filename=download_file) 
        except Exception as e:
            print(e)
            return reports2('failed to download file')
@app.route("/reports2/<string:error>/")
def reports2(error):
    try:
        onlyfiles = [f for f in os.listdir('static/reports2')]
        template = 'reports2.html'
        return render_template(template,files=onlyfiles, error = error)
    except Exception as e:
        print(e)
        return home(e)
    
@app.route("/delete2/<string:name>")
def delete2(name):
    error = 'Document Deleted Successfully'
    try:
        os.remove('static/reports2/'+name)
    except Exception as e:
        print(e)
        error = 'Failed to delete File'
    print('deleting ',name)
    return reports2(error)


@app.route("/logout/")
def logout():
    #session['user'] = None
    return index() 


if __name__ == '__main__':
    app.run(debug=True, port=5000)

    
